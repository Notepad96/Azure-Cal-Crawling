from selenium import webdriver
from selenium.webdriver.support.ui import Select
from pprint import pprint
from openpyxl import Workbook, drawing as draw
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
import requests, time


# WorkBook(파일) 생성
wb = Workbook()
ws = wb.active
ws.title = 'Virtual-Machine'

# Webdriver 사용하여 접속
#driver = webdriver.Chrome('C:\python\chromedriver')
driver = webdriver.Chrome('C:\still\pythonWeb\chromedriver')
driver.get("https://azure.microsoft.com/ko-kr/pricing/calculator")

# 페이지가 로드되는 시간을 위하여 정지
time.sleep(5)

# Virtual-Machine 버튼 클릭하여 추가
vm = driver.find_element_by_xpath('//*[@id="products-picker-panel"]/div[2]/div[2]/div[1]/div/div/div[1]/button')
vm.click()

# VM 추가되는 동안 정지
time.sleep(5)

# 지역 리스트 갖고오기
#region = driver.find_elements_by_css_selector('select[name=region]>option')

currency = ['KRW', 'USD']
cur_text = ['KRW(₩)', 'USD($)']
cur_text2 = ['₩', '$']
cur_size = len(currency) 
region = ['korea-central', 'korea-south']
reg_text = ['한국 중부', '한국 남부']
reg_size = len(region)
operatingSystem = driver.find_elements_by_css_selector('select[name=operatingSystem]>option')


sel_region = Select(driver.find_element_by_xpath('//*[@name="region"]'))
sel_currency = Select(driver.find_element_by_xpath('//*[@class="select currency-dropdown"]'))
sel_os = Select(driver.find_element_by_xpath('//*[@name="operatingSystem"]'))
sel_type = Select(driver.find_element_by_xpath('//*[@name="type"]'))
sel_tier = Select(driver.find_element_by_xpath('//*[@name="tier"]'))


col_text = ["통화", "지역", "운영 체제", "유형", "계층", "인스턴스", "인스턴스2", "인스턴스3"]
col = "ABCDEFGHIJKLMNOPQRSTU"
row_index = 4
col_size = len(col_text)

# 머리글 추가
for i in range(col_size) :
    ws[f'{col[i]}3'] = col_text[i]


hours = driver.find_element_by_name("hours")
hours.clear()
hours.send_keys("1")


sel_currency.select_by_value(currency[0])
sel_region.select_by_value(region[0])
sel_os.select_by_value('windows')
sel_type.select_by_value('sql')
sel_tier.select_by_value('standard')

sel_instance = Select(driver.find_element_by_xpath('//*[@name="size"]'))
sel_instance.select_by_value('ds4v2')


savings = driver.find_elements_by_class_name('savings-option')

print(savings[1].find_element_by_class_name('text-heading5').text[:2])
print(savings[1].find_element_by_css_selector('div[class=total]>span').text)

for i in range(len(savings)) :
    radio_com = savings[i].find_elements_by_css_selector('input')
    for rc in radio_com :
        if rc.is_enabled() :
            rc.click()
            text = driver.find_element_by_css_selector('select[name=size]>option[value=ds4v2]').text
            info = text.split(',')
            print(info[0].split(':')[1].strip())
            print(info[1].strip().split(' ')[0])
            print(info[2].strip().split(' ')[0])
            print(text[text.find(cur_text2[0]):text.find('(')])
            print(f'{col[0]}{row_index}')
            
            ws[f'{col[0]}{row_index}'] = text
            ws[f'{col[1]}{row_index}'] = info[0].split(':')[1].strip()
            ws[f'{col[2]}{row_index}'] = info[1].strip().split(' ')[0]
            ws[f'{col[3]}{row_index}'] = info[2].strip().split(' ')[0]
            ws[f'{col[4]}{row_index}'] = text[text.find(cur_text2[0]):text.find('(')]
        else :
            for i in range(5) :
                ws[f'F{row_index}'] = "X"
        row_index += 1          


wb.save("test.xlsx")


