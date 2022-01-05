from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from pprint import pprint
from openpyxl import Workbook
import time, copy

def check_exists_license():
    try:
        driver.find_element_by_name("license")
    except NoSuchElementException:
        return False
    return True

# WorkBook(파일) 생성
wb = Workbook()
ws = wb.active
ws.title = 'VM-List'

# Webdriver 사용하여 접속
driver = webdriver.Chrome('C:\pythonWeb\chromedriver')
driver.get("https://azure.microsoft.com/ko-kr/pricing/calculator")

# 페이지가 로드되는 시간을 위하여 정지
time.sleep(8)

# Virtual-Machine 버튼 클릭하여 추가
vm = driver.find_element_by_xpath('//*[@id="products-picker-panel"]/div[2]/div[2]/div[1]/div/div/div[1]/button')
vm.click()

# VM 추가되는 동안 정지
time.sleep(5)

# 지역 리스트 갖고오기
#region = driver.find_elements_by_css_selector('select[name=region]>option')


currency = ['KRW', 'USD']
cur_text = ['KRW(₩)', 'USD($)']
cur_text2 = ['₩', '$']
cur_size = len(currency) 
region = ['korea-central', 'korea-south']
reg_text = ['한국 중부', '한국 남부']
reg_size = len(region)
operatingSystem = driver.find_elements_by_css_selector('select[name=operatingSystem]>option')

sel_region = Select(driver.find_element_by_xpath('//*[@name="region"]'))
sel_currency = Select(driver.find_element_by_xpath('//*[@class="select currency-dropdown"]'))
sel_os = Select(driver.find_element_by_xpath('//*[@name="operatingSystem"]'))
sel_type = Select(driver.find_element_by_xpath('//*[@name="type"]'))
sel_tier = Select(driver.find_element_by_xpath('//*[@name="tier"]'))

# 6 + 9 * 3 = 33
col_text = ["통화", "지역", "운영 체제", "유형", "계층", "라이선스", "인스턴스(종량제)", "CPU", "RAM", "임시 스토리지", "가격/시간", "OS-Azure 하이브리드 혜택 지원", "가격/시간", "SW-Azure 하이브리드 혜택 지원", "가격/시간", "인스턴스(1년 예약)", "CPU", "RAM", "임시 스토리지", "가격/시간", "OS-Azure 하이브리드 혜택 지원", "가격", "SW-Azure 하이브리드 혜택 지원", "가격", "인스턴스(3년 예약)", "CPU", "RAM", "임시 스토리지", "가격/시간", "OS-Azure 하이브리드 혜택 지원", "가격", "SW-Azure 하이브리드 혜택 지원", "가격"]
col = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")    # 26
col.extend(["AA", "AB", "AC", "AD", "AE", "AF", "AG"]) # 6
col_size = len(col_text)
index = 4

# 머리글 추가
for i in range(col_size) :
    ws[f'{col[i]}3'] = col_text[i]


# 사용 시간 1 초기화
hours = driver.find_element_by_name("hours")
hours.clear()
hours.send_keys("1")


for cur in range(cur_size) :
    sel_currency.select_by_value(currency[cur])
    for reg in range(reg_size) :
        sel_region.select_by_value(region[reg])
        for osTag in operatingSystem :
            sel_os.select_by_value(osTag.get_attribute('value'))
            tpe = driver.find_elements_by_css_selector('select[name=type]>option')
            for tpTag in tpe :
                sel_type.select_by_value(tpTag.get_attribute('value'))
                tier = driver.find_elements_by_css_selector('select[name=tier]>option')
                for trTag in tier :
                    sel_tier.select_by_value(trTag.get_attribute('value'))
                    if check_exists_license() :
                        sel_license = Select(driver.find_element_by_xpath('//*[@name="license"]'))
                        lic = driver.find_elements_by_css_selector('select[name=license]>option')
                        for licTag in lic :
                            sel_license.select_by_value(licTag.get_attribute('value'))
                            instance = driver.find_elements_by_css_selector('select[name=size]>option')
                            ins_value = list()
                            for va in instance :
                                ins_value.append(va.get_attribute('value'))
                            for ins in ins_value :
                                row = list()
                                savings = driver.find_elements_by_class_name('savings-option')
                                save_len = len(savings)
                                radio_com = savings[0].find_elements_by_css_selector('input')
                                for rc in radio_com :
                                    if rc.is_enabled() :
                                        rc.click()
                                        sel_instance = Select(driver.find_element_by_xpath('//*[@name="size"]'))
                                        sel_instance.select_by_value(ins)
                                        text = driver.find_element_by_css_selector(f'select[name=size]>option[value={ins}').text
                                        row.append(text)
                                        info = text.split(',')
                                        row.append(info[0].split(':')[1].strip())
                                        row.append(info[1].strip().split(' ')[0])
                                        row.append(info[2].strip().split(' ')[0])
                                        row.append(text[text.find(cur_text2[0]):text.find('(')])
                                    else :
                                        row.extend(['-', '-', '-', '-', '-'])
                                    if save_len > 1 : 
                                        if savings[1].find_element_by_class_name('text-heading5').text[:2] == 'OS' :
                                            for k in range(1, save_len) :
                                                radio_what = savings[k].find_elements_by_css_selector('input')
                                                if len(radio_what) > 1 and radio_what[1].is_enabled() :
                                                    row.append('O')
                                                else :
                                                    row.append('X')
                                                row.append(savings[1].find_element_by_css_selector('div[class=total]>span').text)
                                            if save_len == 2 :
                                                row.extend(['-', '-'])
                                        else :
                                            row.extend(['-', '-'])
                                            radio_what = savings[1].find_elements_by_css_selector('input')
                                            if len(radio_what) > 1 and radio_what[1].is_enabled() :
                                                row.append('O')
                                            else :
                                                row.append('X')
                                            row.append(savings[1].find_element_by_css_selector('div[class=total]>span').text)
                                    else :
                                        row.extend(['-', '-', '-', '-'])
                                ws[f'A{index}'] = cur_text[cur]
                                ws[f'B{index}'] = reg_text[reg]
                                ws[f'C{index}'] = osTag.text
                                ws[f'D{index}'] = tpTag.text
                                ws[f'E{index}'] = trTag.text
                                ws[f'F{index}'] = licTag.text
                                for k in range(6, col_size) :
                                    ws[f'{col[k]}{index}'] = row[k-6]
                                index += 1
                    else :
                        instance = driver.find_elements_by_css_selector('select[name=size]>option')
                        ins_value = list()
                        for va in instance :
                            ins_value.append(va.get_attribute('value'))
                        for ins in ins_value :
                            row = list()
                            savings = driver.find_elements_by_class_name('savings-option')
                            save_len = len(savings)
                            radio_com = savings[0].find_elements_by_css_selector('input')
                            for rc in radio_com :
                                if rc.is_enabled() :
                                    rc.click()
                                    sel_instance = Select(driver.find_element_by_xpath('//*[@name="size"]'))
                                    sel_instance.select_by_value(ins)
                                    text = driver.find_element_by_css_selector(f'select[name=size]>option[value={ins}').text
                                    row.append(text)
                                    info = text.split(',')
                                    row.append(info[0].split(':')[1].strip())
                                    row.append(info[1].strip().split(' ')[0])
                                    row.append(info[2].strip().split(' ')[0])
                                    row.append(text[text.find(cur_text2[0]):text.find('(')])
                                else :
                                    row.extend(['-', '-', '-', '-', '-'])
                                if save_len > 1 : 
                                    if savings[1].find_element_by_class_name('text-heading5').text[:2] == 'OS' :
                                        for k in range(1, save_len) :
                                            radio_what = savings[k].find_elements_by_css_selector('input')
                                            if len(radio_what) > 1 and radio_what[1].is_enabled() :
                                                row.append('O')
                                            else :
                                                row.append('X')
                                            row.append(savings[1].find_element_by_css_selector('div[class=total]>span').text)
                                        if save_len == 2 :
                                            row.extend(['-', '-'])
                                    else :
                                        row.extend(['-', '-'])
                                        radio_what = savings[1].find_elements_by_css_selector('input')
                                        if len(radio_what) > 1 and radio_what[1].is_enabled() :
                                            row.append('O')
                                        else :
                                            row.append('X')
                                        row.append(savings[1].find_element_by_css_selector('div[class=total]>span').text)
                                else :
                                    row.extend(['-', '-', '-', '-'])

                            ws[f'A{index}'] = cur_text[cur]
                            ws[f'B{index}'] = reg_text[reg]
                            ws[f'C{index}'] = osTag.text
                            ws[f'D{index}'] = tpTag.text
                            ws[f'E{index}'] = trTag.text
                            ws[f'F{index}'] = '-'
                            for k in range(6, col_size) :
                                ws[f'{col[k]}{index}'] = row[k-6]
                            index += 1


wb.save("Combine.xlsx")




