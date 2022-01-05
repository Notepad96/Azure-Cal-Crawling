from selenium import webdriver
from selenium.webdriver.support.ui import Select
from pprint import pprint
from openpyxl import Workbook, drawing as draw
from openpyxl.styles import Font, Alignment
from openpyxl.styles.borders import Border, Side
import requests, time


# WorkBook(파일) 생성
wb = Workbook()
ws = wb.active
ws.title = 'Virtual-Machine'

# Webdriver 사용하여 접속
driver = webdriver.Chrome('C:\python\chromedriver')
driver.get("https://azure.microsoft.com/ko-kr/pricing/calculator")

# 페이지가 로드되는 시간을 위하여 정지
time.sleep(10)

# Virtual-Machine 버튼 클릭하여 추가
vm = driver.find_element_by_xpath('//*[@id="products-picker-panel"]/div[2]/div[2]/div[1]/div/div/div[1]/button')
vm.click()

# VM 추가되는 동안 정지
time.sleep(5)

# 지역 리스트 갖고오기
#region = driver.find_elements_by_css_selector('select[name=region]>option')

currency = ['KRW', 'USD']
cur_text = ['한국 원(₩)', '미국 달러($)']
cur_size = len(currency) 
region = ['korea-central', 'korea-south']
reg_text = ['한국 중부', '한국 남부']
reg_size = len(region)
operatingSystem = driver.find_elements_by_css_selector('select[name=operatingSystem]>option')

sel_region = Select(driver.find_element_by_xpath('//*[@name="region"]'))
sel_currency = Select(driver.find_element_by_xpath('//*[@class="select currency-dropdown"]'))
sel_os = Select(driver.find_element_by_xpath('//*[@name="operatingSystem"]'))
sel_type = Select(driver.find_element_by_xpath('//*[@name="type"]'))
sel_tier = Select(driver.find_element_by_xpath('//*[@name="tier"]'))
sel_os = Select(driver.find_element_by_xpath('//*[@name="operatingSystem"]'))



driver.find_element_by_css_selector("input[type='radio'][value='three-year']").click()

col_text = ["통화", "지역", "운영 체제", "유형", "계층", "인스턴스"]
index = 1

# 머리글 추가
ws['A1'] = col_text[0]
ws['B1'] = col_text[1]
ws['C1'] = col_text[2]
ws['D1'] = col_text[3]
ws['E1'] = col_text[4]
ws['F1'] = col_text[5]



for cur in range(cur_size) :
    sel_currency.select_by_value(currency[cur])
    for reg in range(reg_size) :
        sel_region.select_by_value(region[reg])
        for osTag in operatingSystem :
            sel_os.select_by_value(osTag.get_attribute('value'))
            tpe = driver.find_elements_by_css_selector('select[name=type]>option')
            for tpTag in tpe :
                sel_type.select_by_value(tpTag.get_attribute('value'))
                tier = driver.find_elements_by_css_selector('select[name=tier]>option')
                for trTag in tier :
                    sel_tier.select_by_value(trTag.get_attribute('value'))
                    instance = driver.find_elements_by_css_selector('select[name=size]>option')
                    for ins in instance :
                        index += 1
                        ws[f'A{index}'] = cur_text[cur]
                        ws[f'B{index}'] = reg_text[reg]
                        ws[f'C{index}'] = osTag.text
                        ws[f'D{index}'] = tpTag.text
                        ws[f'E{index}'] = trTag.text
                        ws[f'F{index}'] = ins.text

"""
for osTag in operatingSystem :
    os = osTag.get_attribute('value')
    ws[f'{col[index]}1'] = os
    sel_os.select_by_value(os)
    tp = driver.find_elements_by_css_selector('select[name=type]>option')
    
    i = 3
    for t in tp :
        ws[f'{col[index]}{i}'] = t.text
        i += 1
    index += 1
"""

wb.save("Azure_Cal3.xlsx")





